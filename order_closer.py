import selenium
from selenium import webdriver
import os
from bs4 import BeautifulSoup
import openpyxl
import datetime
import time
import csv
from Files.file_io import LeanDownloader
from settings import CHROMEDRIVER, DOWNLOAD_DIR, STAR_PAYLOAD, SBW_PAYLOAD

class OrderCloser:
	col_names = ["Client", "Order Id", "Tracking",
				"Date of Shipment", "Product Sku #",
				"Quantity", "Price", "HST"]

	def __init__(self, back_search_depth=50):
		self.open_orders = list()
		self.orders_found = 0
		self.back_search = back_search_depth

	def grab_order_input(self, output_sheet_path):
		path = os.path.abspath(output_sheet_path)
		dir_files = os.listdir(path)
		print("\n")
		for ix, file in enumerate(dir_files):
			print("{} | {}".format(ix, file))

		file_path = dir_files[int(input("\nPick File Number: "))]

		return os.path.join(output_sheet_path, file_path)

	def compile_open_orders(self, file_path):
		wb = openpyxl.load_workbook(file_path)
		open_sheet = wb.active
		open_orders = list()

		# Make list of every open order
		start_row = 2
		end_row = open_sheet.max_row

		for row in range(start_row, end_row + 1):
			
			client = str(open_sheet["BD" + str(row)].value)
			order_id = str(open_sheet["BK" + str(row)].value)
			sku = str(open_sheet["CR" + str(row)].value)
			qty = int(open_sheet["BM" + str(row)].value)
			price = float(open_sheet["CN" + str(row)].value)
			hst = round((qty * price * 0.13), 2)

			self.open_orders.append([client, order_id, sku, qty, price, hst])
			print('Order {} added.'.format(order_id))

		print('\nAll open orders compiled.\n')

	def find_tracking(self, order_history):
		wb = openpyxl.load_workbook(order_history)
		order_sheet = wb.active
		for order in self.open_orders:
			open_po = order[1]
			found = False
			for row in range(order_sheet.max_row, self.back_search, -1):
				lean_po = str(order_sheet["C" + str(row)].value)
				if open_po == lean_po:
					tracking = str(order_sheet["P" + str(row)].value)
					found = True
					if tracking:
						print('Found tracking # for {}'.format(open_po))
						self.orders_found += 1
					else:
						tracking = 'NO TRACKING #'

					shipment_date = str(order_sheet["E" + str(row)].value)[:10]
					order.insert(2, tracking)
					order.insert(3, shipment_date)
					break

	def output_orders(self):
		output_wb = openpyxl.Workbook()
		output_sheet = output_wb.active
		col_width = 20
		for col in range(1, len(self.col_names) + 1):
			col_letter = openpyxl.cell.cell.get_column_letter(col)
			output_sheet.column_dimensions[col_letter].width = col_width

		output_sheet.append(self.col_names)
		for order in self.open_orders:
			output_sheet.append(order)

		print("Found tracking for {} out of {} orders.".format(self.orders_found, len(self.open_orders)))
		
		output_wb.save('closed.xlsx')

if __name__ == '__main__':
	
	orders = OrderCloser()
	file = orders.grab_order_input("Output Sheets")
	orders.compile_open_orders(file)
	dl = LeanDownloader(CHROMEDRIVER)

	if "ORDERS" in file:
		
		dl.login(STAR_PAYLOAD)
		order_sheet = dl.download_all_orders(DOWNLOAD_DIR)
		dl.logout()
		dl.login(SBW_PAYLOAD)
		order_sheet2 = dl.download_all_orders(DOWNLOAD_DIR)
		dl.close()

		orders.find_tracking(order_sheet)
		orders.find_tracking(order_sheet2)
		orders.output_orders()

	elif "STAR" in file:

		dl.login(STAR_PAYLOAD)
		order_sheet = dl.download_all_orders(DOWNLOAD_DIR)
		dl.close()

		orders.find_tracking(order_sheet)
		orders.output_orders()

	elif "SBW" in file:

		dl.login(SBW_PAYLOAD)
		order_sheet = dl.download_all_orders(DOWNLOAD_DIR)
		dl.close()

		orders.find_tracking(order_sheet)
		orders.output_orders()

	else:
		print("Please select an appropriate order file.")