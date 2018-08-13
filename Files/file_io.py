import selenium
from selenium import webdriver
import os
import openpyxl
from bs4 import BeautifulSoup

class LeanDownloader:
	portal_url = "http://4pl.leansupplysolutions.com/portal/login.php"
	index_url = "http://4pl.leansupplysolutions.com/portal/index.php"
	inventory_url = "http://4pl.leansupplysolutions.com/portal/inventory.php"
	order_url = "http://4pl.leansupplysolutions.com/portal/order.php"
	
	def __init__(self, driver_path):
		self.driver = webdriver.Chrome(driver_path)
		self.driver.get(self.portal_url)

	def login(self, payload):
		username, password = payload["username"], payload["userpassword"]
		user = self.driver.find_element_by_name("username")
		user.send_keys(username)

		pw = self.driver.find_element_by_name("userpassword")
		pw.send_keys(password)

		self.driver.find_element_by_id("cmdsave").click()
		print("Login Successful.")

	def download_inventory(self, download_dir):
		self.driver.get(self.inventory_url)
		
		download_before = os.listdir(download_dir)
		downloaded_file = ''
		
		self.driver.find_element_by_id("downloadbtn").click()
		while downloaded_file.endswith('.xls') is False:
			download_after = os.listdir(download_dir)
			for file in download_after:
				if file not in download_before:
					downloaded_file = file

		print("Successfully Downloaded.")
		return os.path.join(download_dir, downloaded_file)

	def download_all_orders(self, download_dir):
		self.driver.get(self.order_url)
		
		download_before = os.listdir(download_dir)
		downloaded_file = ''

		self.driver.find_element_by_css_selector('.btn.btn-warning.dropdown-toggle').click()
		self.driver.find_element_by_id("e0xv8").click()
		self.driver.find_element_by_id("e0xv10").click()

		while downloaded_file.endswith('.xlsx') is False:
			download_after = os.listdir(download_dir)
			for file in download_after:
				if file not in download_before:
					downloaded_file = file

		print("Successfully Downloaded.")
		return os.path.join(download_dir, downloaded_file)


	def logout(self):
		self.driver.get(self.index_url)
		self.driver.find_element_by_id("topmenu_a").click()
		print("Logout Successful.")


	def close(self):
		self.driver.close()
		print("Driver closed.")

def _csv_check(file):
	try:
		if file.endswith('.csv'):
			print('Converting {} to .xlsx format...'.format(file))
			file_name = file.split('.csv')[0]
			wb = openpyxl.Workbook()
			sheet = wb.active

			with open(file, 'r') as f:
				reader = csv.reader(f)
				for r, row in enumerate(reader):
					for c, col in enumerate(row):
							cell = sheet.cell(row=r+1, column=c+1)
							cell.value = col
			return sheet
		else:
			wb = openpyxl.load_workbook(file)
		sheet = wb.active

		return sheet
	except openpyxl.utils.exceptions.InvalidFileException:
		print("{} is not an applicable .csv file.".format(file))

def _convert_xml_to_xlsx(filename, name, inventory_dir):
	file = open(filename, 'r')
	soup = BeautifulSoup(file, 'lxml')
	wb = openpyxl.Workbook()
	sheet = wb.active
	table = []
	for row in soup.find_all("row"):
		new_row = []
		for cell in row.find_all("cell"):
			new_row.append(str(cell.get_text()))
		table.append(new_row)

	for r, row in enumerate(table):
		for c, col in enumerate(row):
			cell = sheet.cell(row=r+1, column=c+1)
			cell.value = col
	output = name + ".xlsx"
	wb.save(os.path.join(inventory_dir, output))

	return os.path.join(inventory_dir, output)

if __name__ == '__main__':
	_csv_check("__init__.py")